## Machine Learning/ Neuronales Netzwerk (MLP) fÃ¼r Voreilstudie 3 ##

# Multilayer Perceptron 
import numpy
import matplotlib.pyplot as plt
from pandas import read_csv
import math
from keras.models import Sequential
from keras.layers import Dense
from openpyxl import load_workbook
import xlsxwriter

# fix random seed for reproducibility
numpy.random.seed(7)

# load the dataset
filenamePar = 'Results.xlsx'
wb = load_workbook(filenamePar, data_only=True)

#defining parameters

X=[]
V2=[]
VE=[]
# get data and save it in lists (here 4 inputs: Stichabnahme (SA), Brammendicke (BD), Temperature (T), Reibkoeffizient (F)
# and one output: Voreilung (VE))
for my_sheet in wb:
    for row in my_sheet.iter_rows():
        if type(row[0].value) is str:  
            if 'yes' in str(row[0].value):
                X.append([row[1].value,row[2].value,row[4].value,row[5].value])
                V2.append(row[7].value)
#calculate the "Voreilung"
for n in range(len(V2)):
    VE.append([abs(2500-V2[n])/2500])
    
# Create a workbook and add a worksheet to print results (errors) in excel-file
X = numpy.array(X)
X=X.astype('float32')
Y = numpy.array(VE)
Y=Y.astype('float32')

# split into train and test sets
trainX, testX = X,X
trainY, testY = Y,Y

# Create a workbook and add a worksheet.
workbook = xlsxwriter.Workbook('ML_results.xlsx')
worksheet = workbook.add_worksheet()

worksheet.write(0,0,'neurons')
worksheet.write(0,1,'layers')
worksheet.write(0,2,'Train Score (MSE)')
worksheet.write(0,3,'Test Score (MSE)')

#define function that creates and excecutes MLP model
def model(n,l,k):
    # create and fit Multilayer Perceptron model
    model = Sequential()
    for i in range(l):
        model.add(Dense(n, input_dim=4, activation='relu'))
    model.add(Dense(1))
    model.compile(loss='mean_squared_error', optimizer='adam')
    model.fit(trainX, trainY, epochs=200, batch_size=2, verbose=2)
    
    # Estimate model performance
    trainScore = model.evaluate(trainX, trainY, verbose=0)
    print('Train Score: %g MSE (%.2f RMSE)' % (trainScore, math.sqrt(trainScore)))
    testScore = model.evaluate(testX, testY, verbose=0)
    print('Test Score: %g MSE (%.2f RMSE)' % (testScore, math.sqrt(testScore)))
    
    # generate predictions for training
    YtrainPredict = model.predict(trainX)
    YtestPredict = model.predict(testX)
    
    
    # plot baseline and predictions
    fig=plt.figure()
    plt.plot(Y)
    plt.plot(YtrainPredict,'r')
    plt.plot(YtestPredict,'g')
    fig.savefig('ML_n'+str(n)+'_l'+str(l)+'.png')

    #write errors in excel-file
    worksheet.write(k,0,n)
    worksheet.write(k,1,l)
    worksheet.write(k,2,trainScore)
    worksheet.write(k,3,testScore)
    
#excecute for different neuron (n) and layer (l) numbers
k=1
for n in range(8,21,2):
    for l in range(1,5):
        model(n,l,k)
        k+=1

workbook.close()
