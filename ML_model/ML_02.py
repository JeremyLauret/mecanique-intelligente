## Machine Learning/ Neuronales Netzwerk (MLP) für Voreilstudie 2 ##

# Multilayer Perceptron 
import numpy
import matplotlib.pyplot as plt
from pandas import read_csv
import math
from keras.models import Sequential
from keras.layers import Dense
from openpyxl import load_workbook
import xlsxwriter

# fix random seed for reproducibility
numpy.random.seed(7)

# load the dataset
filenamePar = 'Results.xlsx'
wb = load_workbook(filenamePar, data_only=True)

#defining parameters

X=[]
V2=[]
VE=[]

#get data and save it in lists (here two inputs: Stichabnahme (SA), Brammendicke (BD) and one output: Voreilung (VE))
for my_sheet in wb:
    for row in my_sheet.iter_rows():
        if type(row[0].value) is str:  
            if 'yes' in str(row[0].value):
                X.append([row[1].value,row[2].value])
                V2.append(row[5].value)
#calculate the "Voreilung"
for n in range(len(V2)):
    VE.append([abs(2500-V2[n])/2500])
    
# insert data in input-array X and output-array Y
X = numpy.array(X)
X=X.astype('float32')
Y = numpy.array(VE)
Y=Y.astype('float32')

# split into train and test sets
trainX, testX = numpy.concatenate((X[0:250,:],X[350:750,:],X[850:1250,:],X[1350:1750,:],X[1850:2250,:],X[2350:2750,:],X[2850:3250,:],X[3350:3750,:],X[3850:len(X),:]), axis=0),  numpy.concatenate((X[250:350,:],X[750:850,:],X[1250:1350,:],X[1750:1850,:],X[2250:2350,:],X[2750:2850,:],X[3250:3350,:],X[3750:3850,:]),axis=0)
trainY, testY = numpy.concatenate((Y[0:250,:],Y[350:750,:],Y[850:1250,:],Y[1350:1750,:],Y[1850:2250,:],Y[2350:2750,:],Y[2850:3250,:],Y[3350:3750,:],Y[3850:len(X),:]), axis=0),  numpy.concatenate((Y[250:350,:],Y[750:850,:],Y[1250:1350,:],Y[1750:1850,:],Y[2250:2350,:],Y[2750:2850,:],Y[3250:3350,:],Y[3750:3850,:]),axis=0)


# Create a workbook and add a worksheet to print results (errors) in excel-file
workbook = xlsxwriter.Workbook('ML_results.xlsx')
worksheet = workbook.add_worksheet()

worksheet.write(0,0,'neurons')
worksheet.write(0,1,'layers')
worksheet.write(0,2,'Train Score (MSE)')
worksheet.write(0,3,'Test Score (MSE)')

#define function that creates and excecutes MLP model
def model(n,l,k):
    # create and fit Multilayer Perceptron model
    model = Sequential()
    for i in range(l):
        model.add(Dense(n, input_dim=2, activation='relu'))
    model.add(Dense(1))
    model.compile(loss='mean_squared_error', optimizer='adam')
    model.fit(trainX, trainY, epochs=200, batch_size=2, verbose=2)
    
    # Estimate model performance
    trainScore = model.evaluate(trainX, trainY, verbose=0)
    print('Train Score: %g MSE (%.2f RMSE)' % (trainScore, math.sqrt(trainScore)))
    testScore = model.evaluate(testX, testY, verbose=0)
    print('Test Score: %g MSE (%.2f RMSE)' % (testScore, math.sqrt(testScore)))
    
    # generate predictions for training
    YtrainPredict = model.predict(trainX)
    YtestPredict = model.predict(testX)
    
    # shift train predictions for plotting
    trainPredictPlot1 = numpy.empty_like(X)
    trainPredictPlot1[:, :] = numpy.nan
    trainPredictPlot1[0:250, :] = YtrainPredict[0:250,:]
    
    trainPredictPlot2 = numpy.empty_like(X)
    trainPredictPlot2[:, :] = numpy.nan
    trainPredictPlot2[350:750, :] = YtrainPredict[250:650,:]
    
    trainPredictPlot3 = numpy.empty_like(X)
    trainPredictPlot3[:, :] = numpy.nan
    trainPredictPlot3[850:1250, :] = YtrainPredict[650:1050,:]
    
    trainPredictPlot4 = numpy.empty_like(X)
    trainPredictPlot4[:, :] = numpy.nan
    trainPredictPlot4[1350:1750, :] = YtrainPredict[1050:1450,:]
    
    trainPredictPlot5 = numpy.empty_like(X)
    trainPredictPlot5[:, :] = numpy.nan
    trainPredictPlot5[1850:2250, :] = YtrainPredict[1450:1850,:]
    
    trainPredictPlot6 = numpy.empty_like(X)
    trainPredictPlot6[:, :] = numpy.nan
    trainPredictPlot6[2350:2750, :] = YtrainPredict[1850:2250,:]
    
    trainPredictPlot7 = numpy.empty_like(X)
    trainPredictPlot7[:, :] = numpy.nan
    trainPredictPlot7[2850:3250, :] = YtrainPredict[2250:2650,:]
    
    trainPredictPlot8 = numpy.empty_like(X)
    trainPredictPlot8[:, :] = numpy.nan
    trainPredictPlot8[3350:3750, :] = YtrainPredict[2650:3050,:]
    
    trainPredictPlot9 = numpy.empty_like(X)
    trainPredictPlot9[:, :] = numpy.nan
    trainPredictPlot9[3850:len(X), :] = YtrainPredict[3050:len(YtrainPredict),:]
    
    # shift test predictions for plotting
    testPredictPlot1 = numpy.empty_like(X)
    testPredictPlot1[:, :] = numpy.nan
    testPredictPlot1[250:350, :] = YtestPredict[0:100,:]
    
    testPredictPlot2 = numpy.empty_like(X)
    testPredictPlot2[:, :] = numpy.nan
    testPredictPlot2[750:850, :] = YtestPredict[100:200,:]
    
    testPredictPlot3 = numpy.empty_like(X)
    testPredictPlot3[:, :] = numpy.nan
    testPredictPlot3[1250:1350, :] = YtestPredict[200:300,:]
    
    testPredictPlot4 = numpy.empty_like(X)
    testPredictPlot4[:, :] = numpy.nan
    testPredictPlot4[1750:1850, :] = YtestPredict[300:400,:]
    
    testPredictPlot5 = numpy.empty_like(X)
    testPredictPlot5[:, :] = numpy.nan
    testPredictPlot5[2250:2350, :] = YtestPredict[400:500,:]
    
    testPredictPlot6 = numpy.empty_like(X)
    testPredictPlot6[:, :] = numpy.nan
    testPredictPlot6[2750:2850, :] = YtestPredict[500:600,:]
    
    testPredictPlot7 = numpy.empty_like(X)
    testPredictPlot7[:, :] = numpy.nan
    testPredictPlot7[3250:3350, :] = YtestPredict[600:700,:]
    
    testPredictPlot8 = numpy.empty_like(X)
    testPredictPlot8[:, :] = numpy.nan
    testPredictPlot8[3750:3850, :] = YtestPredict[700:800,:]
    
    # plot baseline and predictions
    fig=plt.figure()
    
    plt.plot(Y)
    
    plt.plot(trainPredictPlot1,'r')
    plt.plot(trainPredictPlot2,'r')
    plt.plot(trainPredictPlot3,'r')
    plt.plot(trainPredictPlot4,'r')
    plt.plot(trainPredictPlot5,'r')
    plt.plot(trainPredictPlot6,'r')
    plt.plot(trainPredictPlot7,'r')
    plt.plot(trainPredictPlot8,'r')
    plt.plot(trainPredictPlot9,'r')
    
    plt.plot(testPredictPlot1,'g')
    plt.plot(testPredictPlot2,'g')
    plt.plot(testPredictPlot3,'g')
    plt.plot(testPredictPlot4,'g')
    plt.plot(testPredictPlot5,'g')
    plt.plot(testPredictPlot6,'g')
    plt.plot(testPredictPlot7,'g')
    plt.plot(testPredictPlot8,'g')
    
    fig.savefig('ML_n'+str(n)+'_l'+str(l)+'.png')

    #write errors in excel-file
    worksheet.write(k,0,n)
    worksheet.write(k,1,l)
    worksheet.write(k,2,trainScore)
    worksheet.write(k,3,testScore)
    
#excecute for different neuron (n) and layer (l) numbers
k=1
for n in range(8,21,2):
    for l in range(1,5):
        model(n,l,k)
        k+=1

workbook.close()
