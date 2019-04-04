## Machine Learning/ Neuronales Netzwerk (MLP) für Voreilstudie 1 ##

# Import necessary libaries and functions
import numpy
import matplotlib.pyplot as plt
import math
from keras.models import Sequential
from keras.layers import Dense
from openpyxl import load_workbook
import xlsxwriter

# fix random seed for reproducibility
numpy.random.seed(7)

# load the dataset
filenamePar = 'Results.xlsx'
wb = load_workbook(filenamePar, data_only=True)

#defining parameters
SA=[]
V2=[]
VE=[]

#get data and save it in lists (here only one input: Stichabnahme (SA) and one output: Voreilung (VE))
for my_sheet in wb:
    for row in my_sheet.iter_rows():
        if type(row[0].value) is str:  
            if 'yes' in str(row[0].value): 
                SA.append([row[2].value])
                V2.append(row[5].value)
#calculate the "Voreilung"
for n in range(len(V2)):
    VE.append([abs(2500-V2[n])/2500])

# insert data in input-array X and output-array Y
X = numpy.array(SA)
X=X.astype('float32')
Y = numpy.array(VE)
Y=Y.astype('float32')

# split into train and test sets
trainX, testX = numpy.concatenate((X[0:200,:],X[500:len(X),:]), axis=0), X[200:500,:]
trainY, testY = numpy.concatenate((Y[0:200,:],Y[500:len(X),:]), axis=0), Y[200:500,:]

# Create a workbook and add a worksheet to print results (errors) in excel-file
workbook = xlsxwriter.Workbook('ML_results.xlsx')
worksheet = workbook.add_worksheet()

worksheet.write(0,0,'neurons')
worksheet.write(0,1,'layers')
worksheet.write(0,2,'Train Score (MSE)')
worksheet.write(0,3,'Test Score (MSE)')

#define function that creates and excecutes MLP model
def model(n,l,k):
    # create and fit Multilayer Perceptron model
    model = Sequential()
    for i in range(l):
        model.add(Dense(n, input_dim=1, activation='relu'))
    model.add(Dense(1))
    model.compile(loss='mean_squared_error', optimizer='adam')
    model.fit(trainX, trainY, epochs=200, batch_size=2, verbose=2)
    
    # Estimate model performance
    trainScore = model.evaluate(trainX, trainY, verbose=0)
    print('Train Score: %g MSE (%.2f RMSE)' % (trainScore, math.sqrt(trainScore)))
    testScore = model.evaluate(testX, testY, verbose=0)
    print('Test Score: %g MSE (%.2f RMSE)' % (testScore, math.sqrt(testScore)))
    
    # generate predictions for training
    YtrainPredict = model.predict(trainX)
    YtestPredict = model.predict(testX)
    
    # shift train predictions for plotting
    trainPredictPlot1 = numpy.empty_like(X)
    trainPredictPlot1[:, :] = numpy.nan
    trainPredictPlot1[0:200, :] = YtrainPredict[0:200,:]
    
    trainPredictPlot2 = numpy.empty_like(X)
    trainPredictPlot2[:, :] = numpy.nan
    trainPredictPlot2[500:len(X), :] = YtrainPredict[200:len(YtrainPredict),:]
    
    # shift test predictions for plotting
    testPredictPlot = numpy.empty_like(X)
    testPredictPlot[:, :] = numpy.nan
    testPredictPlot[200:500, :] = YtestPredict
    
    # plot baseline and predictions and save figure
    fig=plt.figure()
    plt.plot(Y)
    plt.plot(trainPredictPlot1,'r')
    plt.plot(trainPredictPlot2,'r')
    plt.plot(testPredictPlot, 'g')
    fig.savefig('ML_n'+str(n)+'_l'+str(l)+'.png')
    
    #write errors in excel-file
    worksheet.write(k,0,n)
    worksheet.write(k,1,l)
    worksheet.write(k,2,trainScore)
    worksheet.write(k,3,testScore)

#excecute for different neuron (n) and layer (l) numbers
k=1
for n in range(8,21,2):
    for l in range(1,5):
        model(n,l,k)
        k+=1

workbook.close()